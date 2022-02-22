# -*- encoding: utf-8 -*-
"""
@File    :   file_processing.py
@Time    :   2022/02/17 11:51:44
@Author  :   Yimin Jing
@Contact :   jingym3@lenovo.com
@Desc    :   文件的读取和写入
"""

import os
import json
import glob
import yaml
import pandas as pd
from openpyxl import Workbook, load_workbook

    
def load_single_json(in_path, prefix=None):
    """
    @Desc: 加载Json文件，整个Json文件是一个python字典
    @Args: 
        in_path: Json文件路径
        prefix: 该字典的某个字段 (key，键)
    @Returns: 
        python字典或列表
    """
    with open(in_path, "r", encoding="utf-8") as f:
        if prefix:
            data = json.load(f)[prefix]
        else:
            data = json.load(f)
    return data
    

def load_multi_json(in_path):
    """
    @Desc: 加载Json文件，一行是一个python字典
    @Args: 
        in_path: Json文件路径
    @Returns:
        python列表
    """
    with open(in_path, "r", encoding="utf-8") as f:
        data = [json.loads(line.strip()) for line in f]
    return data
    

def write_json(data, out_path):
    """
    @Desc: 将python字典或字典列表写入Json文件
    @Args: 
        data: python字典或者字典列表
    @Returns: 
        输出json文件
    """
    with open(out_path, "w", encoding="utf-8") as f:
        if isinstance(data, list):
            for ele in data:
                if isinstance(ele, dict):
                    f.write("{}\n".format(json.dumps(ele, ensure_ascii=False)))
        elif isinstance(data, dict):
            f.write(json.dumps(data, ensure_ascii=False, indent=4))
        else:
            raise Exception("写入json数据的文件格式不正确")
    
    
def output_filenames(in_directory, file_type):
    """
    @Desc: 输出给定目录里的所有指定文件类型的绝对路径
    @Args: 
        in_directory: 给定目录
        file_type: 指定文件类型
    @Returns: 
        路径List
    """
    specific_directory = os.path.join(in_directory, "*.{}".format(file_type))
    return glob.glob(specific_directory)
    

def load_single_txt(in_path, keep_enter=False):
    """
    @Desc: 加载单个纯文本文件
    @Args: 
        in_path: 纯文本文件路径
        keep_enter: 用来确定是否保留每一行的换行符"\n"
    @Returns: 
        python列表
    """
    data = []
    with open(in_path, "r", encoding="utf-8") as f:
        for line in f:
            if keep_enter:
                data.append(line)
            else:
                if line.strip():
                    data.append(line.strip())
    return data


def load_multi_txt(in_directory, file_type, keep_enter=False):
    """
    @Desc: 加载给定目录里的所有纯文本文件
    @Args: 
        in_directory: 给定目录
        file_type: 文件类型
        keep_enter: 是否保留换行符
    @Returns: 
        python列表
    """
    filenames = output_filenames(in_directory, file_type)
    data = []
    for filename in filenames:
        res = load_single_txt(filename, keep_enter=keep_enter)
        data.extend(res)
    return data


def write_txt(data, out_path):
    """
    @Desc: 将python列表数据写入纯文本文件
    @Args: 
        data: ["今天天气晴朗", "适合爬山"]
        out_path: 输出的额纯文本文件路径
    @Returns: 
    """
    with open(out_path, "w", encoding="utf-8") as f:
        for ele in data:
            f.write(ele.strip() + "\n")


def load_yaml(in_path):
    """
    @Desc: 加载yaml参数配置文件
    @Args: 
        in_path: yaml文件路径 
    @Returns: 
        python字典
    """
    with open(in_path, "r", encoding="utf-8") as f:
        data = yaml.load(f, Loader=yaml.FullLoader)
    return data


def write_excel(data, out_path, sheet_name="Sheet"):

    """
    @Desc: 将数据写入excel文件
    @Args: 
        data: {"列名1":[ele1, ele2], "列名2": [ele1, ele2]}
        out_path: *.xlsx
        sheet_name: 表单名
    @Returns: 
        excel文件
    """
    if os.path.exists(out_path):
        book = load_workbook(out_path)
        writer = pd.ExcelWriter(out_path, engine="openpyxl", mode="a")
        writer.book = book
    else:
        book = Workbook()
        writer = pd.ExcelWriter(out_path, engine="openpyxl", mode="w")
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()
    
